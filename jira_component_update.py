#!/usr/bin/env python3
# -*- coding: utf-8-unix -*-

from jira import JIRA, JIRAError
import sys, itertools

class jira_component_tools():
    def __init__(self,user,pwd,project,uri, api = r"/rest/api/2/search" ):
        self.user = user
        self.pwd = pwd
        self.project = project
        self.uri = uri
        self.api = api

    def log_in(self):
        try:
            print('Authenticating user', self.user, 'to project', self.project, 'on uri', self.uri)
            
            options = {'server': self.uri}
            self.jira = JIRA(options, basic_auth=(self.user, self.pwd))
            
            print('Authentication success\n')
            
        except JIRAError as e:
            print (datetime.datetime.now(),">>", e.status_code, e.text)
            exit()

    def __get_component_dict(self):
        try:
            component_dict = {} # allows getting id component by component name    
            component_in_proj = self.jira.project_components(self.project)
            
            for component in component_in_proj:        
                component_dict[component.name] = component.id
                
            return component_dict
            
        except JIRAError as e:
            print (datetime.datetime.now(),">>", e.status_code, e.text)
            exit()
            
    def load_components(self,component_dict):    
        spinner = itertools.cycle(['-', '/', '|', '\\'])
        
        jira_component_dump = self.__get_component_dict()    
        not_added = 0
        total_component_count = 0
        
        sys.stdout.write('Processing components: ')
        sys.stdout.flush()
        
        try:
            for key in jira_component_dump:
                desc = jira_component_dump[key]
                          
                if key in jira_component_dump:
                    not_added += 1
                    total_component_count += 1
                    
                else:
                    total_component_count += 1   
                    self.jira.create_component(name, self.project, description=desc, leadUserName=None, assigneeType=None, isAssigneeTypeValid=False)
        
                sys.stdout.write(next(spinner))
                sys.stdout.flush()
                sys.stdout.write('\b')
            
            print('\n')
            print(total_component_count, 'components processed')
            print(not_added, 'already in database', total_component_count - not_added, 'components added to jira database')
            
        except JIRAError as e:
            print (datetime.datetime.now(),">>", e.status_code, e.text)
            exit()
        
if __name__ == "__main__":    

    ### EXAMPLE: LOAD COMPONENTS FROM EXCEL WORKBOOK ###
    #
    #   Format: 
    #       Column A: NAME
    #       Column B: DESCRIPTION
    #    
    from openpyxl import load_workbook
    wb = load_workbook( filename = "components.xlsx" )
    ws = wb['Sheet1']
    i = 2
    component_dict = {}        
    while (ws['A'+str(i)].value):        
        try:
            name = str(ws['A'+str(i)].value)
            desc = str(ws['B'+str(i)].value)              
            component_dict[name] = desc
        except:
            pass
        i += 1
    #
    ###################################################
    
    jira_url = r"https://YOURURL.atlassian.net"    
    auth_user = "YOURUSERNAME"
    auth_pwd = "YOURPASSWORD"
    project = "YOURPROJECT"
    
    j=jira_component_tools(auth_user,auth_pwd,project,jira_url)
    j.log_in()
    j.load_components(component_dict)